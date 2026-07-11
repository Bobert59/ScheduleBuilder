from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook


def make_history_workbook(
    path: Path,
    start: date,
    days: int,
    rows: dict[str, dict[date, str]],
    *,
    split_headers: bool = False,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"
    dates = [start + timedelta(days=offset) for offset in range(days)]
    if split_headers:
        for column, day in enumerate(dates, start=2):
            sheet.cell(1, column, f"{day:%a}")
            sheet.cell(2, column, f"{day:%b %d}")
        first_doctor_row = 3
    else:
        sheet.cell(1, 1, "Doctor")
        for column, day in enumerate(dates, start=2):
            sheet.cell(1, column, f"{day:%a}\n{day:%b %d}")
        first_doctor_row = 2
    for row, (name, assignments) in enumerate(rows.items(), start=first_doctor_row):
        sheet.cell(row, 1, name)
        for column, day in enumerate(dates, start=2):
            sheet.cell(row, column, assignments.get(day))
    workbook.save(path)
    return path
