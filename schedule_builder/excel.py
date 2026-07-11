from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .domain import (
    SHIFT_HOURS,
    SHIFT_NAMES,
    DoctorMode,
    HistorySchedule,
    ScheduleConfig,
    ScheduleResult,
)
from .rules import unavailable_dates_for


HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
SCHEDULE_HEADER_FILL = PatternFill("solid", fgColor="FFE8E8E8")
UNAVAILABLE_FILL = PatternFill("solid", fgColor="FFF2CEEF")
PROTECTED_FILL = PatternFill("solid", fgColor="FFDAF2D0")
OPEN_FILL = PatternFill("solid", fgColor="FFFFFF00")
THIN_GRAY = Side(style="thin", color="FFD9E1F2")
THIN_BLACK = Side(style="thin", color="FF000000")
GRID_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def _open_rows(result: ScheduleResult, dates: tuple[date, ...]) -> list[tuple[str, dict[date, str]]]:
    maximum = max((len(result.open_shifts.get(day, ())) for day in dates), default=0)
    rows: list[tuple[str, dict[date, str]]] = []
    for row_index in range(maximum):
        name = "OPEN" if row_index == 0 else f"OPEN {row_index + 1}"
        assignments = {
            day: result.open_shifts[day][row_index]
            for day in dates
            if len(result.open_shifts.get(day, ())) > row_index
        }
        rows.append((name, assignments))
    return rows


def write_schedule_workbook(
    path: str | Path,
    config: ScheduleConfig,
    history: HistorySchedule,
    result: ScheduleResult,
) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "Schedule"
    dates = config.dates

    corner = schedule.cell(row=1, column=1)
    corner.fill = SCHEDULE_HEADER_FILL
    corner.font = Font(name="Aptos Narrow", size=11, bold=True, color="FF000000")
    for column, day in enumerate(dates, start=2):
        cell = schedule.cell(row=1, column=column, value=f"{day:%a}\n  {day:%b %d}")
        cell.fill = SCHEDULE_HEADER_FILL
        cell.font = Font(name="Aptos Narrow", size=11, bold=True, color="FF000000")

    rows = [(doctor.name, result.assignments[doctor.name]) for doctor in config.doctors]
    rows.extend(_open_rows(result, dates))
    doctor_by_name = {doctor.name: doctor for doctor in config.doctors}
    for row_number, (name, assignments) in enumerate(rows, start=2):
        name_cell = schedule.cell(row=row_number, column=1, value=name)
        name_cell.font = Font(name="Aptos Narrow", size=11, bold=True, color="FF000000")
        doctor = doctor_by_name.get(name)
        unavailable = unavailable_dates_for(doctor, dates) if doctor else set()
        for column, day in enumerate(dates, start=2):
            cell = schedule.cell(row=row_number, column=column, value=assignments.get(day))
            cell.font = Font(name="Aptos Narrow", size=11, color="FF000000")
            if name.upper().startswith("OPEN") and cell.value:
                cell.fill = OPEN_FILL
            elif doctor and day in doctor.assignments:
                cell.fill = PROTECTED_FILL
            elif day in unavailable:
                cell.fill = UNAVAILABLE_FILL

    for row in range(1, schedule.max_row + 1):
        for column in range(1, schedule.max_column + 1):
            cell = schedule.cell(row=row, column=column)
            left = THIN_BLACK if column == 1 else None
            right = THIN_BLACK if column == 1 or dates[column - 2].weekday() == 6 else None
            top = THIN_BLACK if row == 1 and column == 1 else None
            bottom = THIN_BLACK if row == schedule.max_row or (row == 1 and column == 1) else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    summary = workbook.create_sheet("Summary")
    headers = [
        "Doctor",
        "Mode",
        "# 8-6",
        "# 8-8",
        "# 2-12",
        "# O/N",
        "Total Shifts",
        "Days Off",
        "Total Hours",
        "Target Hours",
        "Hours/Week",
        "Weekend Shifts",
        "History O/N",
        "Combined O/N",
    ]
    summary.append(headers)
    weeks = len(dates) / 7
    for doctor in config.doctors:
        assignments = result.assignments[doctor.name]
        counts = {shift: sum(value == shift for value in assignments.values()) for shift in SHIFT_NAMES}
        total_shifts = sum(counts.values())
        total_hours = sum(SHIFT_HOURS[shift] * count for shift, count in counts.items())
        weekend_shifts = sum(day.weekday() >= 5 for day in assignments)
        history_overnights = result.history_overnights[doctor.name]
        summary.append(
            [
                doctor.name,
                doctor.mode.value.title(),
                counts["8-6"],
                counts["8-8"],
                counts["2-12"],
                counts["O/N"],
                total_shifts,
                len(dates) - total_shifts,
                total_hours,
                result.target_hours[doctor.name],
                round(total_hours / weeks, 1),
                weekend_shifts,
                history_overnights,
                history_overnights + counts["O/N"],
            ]
        )
    if result.open_shifts:
        open_counts = {shift: 0 for shift in SHIFT_NAMES}
        for shifts in result.open_shifts.values():
            for shift in shifts:
                open_counts[shift] += 1
        total = sum(open_counts.values())
        total_hours = sum(SHIFT_HOURS[shift] * count for shift, count in open_counts.items())
        summary.append(
            [
                "OPEN",
                "Automatic",
                open_counts["8-6"],
                open_counts["8-8"],
                open_counts["2-12"],
                open_counts["O/N"],
                total,
                "",
                total_hours,
                "",
                "",
                sum(
                    len(shifts)
                    for day, shifts in result.open_shifts.items()
                    if day.weekday() >= 5
                ),
                "",
                "",
            ]
        )
    _style_table(summary)

    details = workbook.create_sheet("Run Details")
    details.append(["Schedule Builder", "2.0"])
    details.append(["Schedule window", f"{config.start:%Y-%m-%d} to {config.end:%Y-%m-%d}"])
    details.append(["History workbook", str(history.source)])
    details.append(["History window", f"{history.dates[0]:%Y-%m-%d} to {history.dates[-1]:%Y-%m-%d}"])
    details.append(["History days imported", len(history.dates)])
    details.append([])
    details.append(["Optimization phase", "Status", "Objective", "Wall time (seconds)"])
    for report in result.phase_reports:
        details.append([report.name, report.status, report.objective, round(report.wall_time_seconds, 3)])
    details.append([])
    details.append(["Legend", "Meaning"])
    details.append(["Green", "Fixed or prescribed assignment"])
    details.append(["Purple", "Unavailable date"])
    details.append(["Yellow", "Automatic OPEN shift"])
    details.append(["Gray header", "Schedule date"])
    details.append([])
    details.append(["OPEN priority (first used)", "Weekday 8-6, weekday 8-8, weekday 2-12, weekend 8-6, weekend 8-8, weekend 2-12, O/N"])
    _style_table(details, header_row=7)
    details.column_dimensions["A"].width = 28
    details.column_dimensions["B"].width = 95
    details.column_dimensions["C"].width = 18
    details.column_dimensions["D"].width = 20

    workbook.save(output)
    return output


def _style_table(sheet, header_row: int = 1) -> None:
    for cell in sheet[header_row]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = Font(color="FFFFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = GRID_BORDER
                if cell.row != header_row:
                    cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.sheet_view.showGridLines = False
    for column in range(1, sheet.max_column + 1):
        values = [len(str(sheet.cell(row=row, column=column).value or "")) for row in range(1, sheet.max_row + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(values, default=8) + 2, 28)
