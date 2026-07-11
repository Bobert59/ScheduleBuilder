from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from .domain import SHIFT_NAMES, HistorySchedule
from .errors import HistoryFormatError


_HEADER_RE = re.compile(
    r"^(?:(?P<weekday>[A-Za-z]{3,9})\s+)?(?P<month>[A-Za-z]{3,9})\s+(?P<day>\d{1,2})$"
)
_WEEKDAYS = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}


def _header_parts(value: object) -> tuple[int, int, int | None]:
    if isinstance(value, datetime):
        return value.month, value.day, value.weekday()
    if isinstance(value, date):
        return value.month, value.day, value.weekday()
    text = " ".join(str(value).replace("\n", " ").split())
    match = _HEADER_RE.match(text)
    if not match:
        raise HistoryFormatError(f"Unrecognized schedule date header: {value!r}.")
    try:
        month = datetime.strptime(match.group("month")[:3], "%b").month
    except ValueError as exc:
        raise HistoryFormatError(f"Unrecognized month in schedule header: {value!r}.") from exc
    weekday_text = match.group("weekday")
    weekday = None
    if weekday_text:
        weekday = _WEEKDAYS.get(weekday_text[:3].lower())
        if weekday is None:
            raise HistoryFormatError(f"Unrecognized weekday in schedule header: {value!r}.")
    return month, int(match.group("day")), weekday


def _shift_name(value: object) -> str | None:
    """Normalize shift cells, including labels Excel automatically converted to dates."""
    if value in (None, ""):
        return None
    if isinstance(value, (datetime, date)):
        candidate = f"{value.month}-{value.day}"
        return candidate if candidate in SHIFT_NAMES else None
    candidate = str(value).strip()
    return candidate if candidate in SHIFT_NAMES else None


def _find_date_header(sheet) -> tuple[int, list[int]]:
    """Find a row containing contiguous month/day headers, with or without weekdays."""
    for row in range(1, min(sheet.max_row, 10) + 1):
        columns = [
            column
            for column in range(2, sheet.max_column + 1)
            if sheet.cell(row=row, column=column).value not in (None, "")
        ]
        if not columns or columns != list(range(2, 2 + len(columns))):
            continue
        try:
            for column in columns:
                _header_parts(sheet.cell(row=row, column=column).value)
        except HistoryFormatError:
            continue
        return row, columns
    raise HistoryFormatError(
        "The Schedule sheet has no recognizable date-header row. Dates may be written as "
        "'Mon\\nAug 24' in one row or as weekday and month/day in two separate rows."
    )


def _weekday_above(sheet, header_row: int, column: int) -> int | None:
    if header_row <= 1:
        return None
    raw_value = sheet.cell(row=header_row - 1, column=column).value
    if raw_value in (None, ""):
        return None
    text = str(raw_value).strip()[:3].lower()
    return _WEEKDAYS.get(text)


def read_history_workbook(path: str | Path, expected_end: date) -> HistorySchedule:
    source = Path(path)
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise HistoryFormatError(f"Could not open history workbook {source}: {exc}") from exc
    try:
        if "Schedule" not in workbook.sheetnames:
            raise HistoryFormatError("The history workbook must contain a sheet named 'Schedule'.")
        sheet = workbook["Schedule"]
        date_header_row, date_columns = _find_date_header(sheet)

        first_day = expected_end - timedelta(days=len(date_columns) - 1)
        dates = tuple(first_day + timedelta(days=i) for i in range(len(date_columns)))
        for column, expected in zip(date_columns, dates, strict=True):
            raw_header = sheet.cell(row=date_header_row, column=column).value
            month, day, weekday = _header_parts(raw_header)
            if weekday is None:
                weekday = _weekday_above(sheet, date_header_row, column)
            if (month, day) != (expected.month, expected.day) or (
                weekday is not None and weekday != expected.weekday()
            ):
                raise HistoryFormatError(
                    f"History date mismatch in column {column}: expected {expected:%a %b %d}, "
                    f"found {raw_header!r}. The workbook must end on "
                    f"{expected_end:%Y-%m-%d}."
                )

        assignments: dict[str, dict[date, str]] = {}
        open_lists: dict[date, list[str]] = {day: [] for day in dates}
        seen: set[str] = set()
        for row in range(date_header_row + 1, sheet.max_row + 1):
            raw_name = sheet.cell(row=row, column=1).value
            if raw_name in (None, ""):
                continue
            name = str(raw_name).strip()
            is_open = name.upper().startswith("OPEN")
            key = name.casefold()
            if key in seen and not is_open:
                raise HistoryFormatError(f"Doctor {name!r} appears more than once in the Schedule sheet.")
            seen.add(key)
            doctor_assignments: dict[date, str] = {}
            for column, day in zip(date_columns, dates, strict=True):
                raw_shift = sheet.cell(row=row, column=column).value
                if raw_shift in (None, ""):
                    continue
                shift = _shift_name(raw_shift)
                if shift is None:
                    raise HistoryFormatError(
                        f"Unknown shift {raw_shift!r} for {name} on {day:%Y-%m-%d}."
                    )
                if is_open:
                    open_lists[day].append(shift)
                else:
                    doctor_assignments[day] = shift
            if not is_open:
                assignments[name] = doctor_assignments

        return HistorySchedule(
            source=source,
            dates=dates,
            assignments=assignments,
            open_shifts={day: tuple(shifts) for day, shifts in open_lists.items() if shifts},
        )
    finally:
        workbook.close()
